# -*- codding:utf-8 -*-


from openpyxl import Workbook
from openpyxl import load_workbook
import time
import os

code_list= [{"KlTm": 1522070100,"TrdDy": "20180327","OpPri": 3354,"PreClPri": 3363,"HiPri": 3378,"LoPri": 3353},
            {"KlTm": 1522071000,"TrdDy": "20180327","OpPri": 3373,"PreClPri": 3363,"HiPri": 3376,"LoPri": 3368},
            {"KlTm": 1522071900,"TrdDy": "20180327","OpPri": 3370,"PreClPri": 3363,"HiPri": 3376,"LoPri": 3364}]
filename="E:\JupyterNotebook\读写xlsx.xlsx"

data_keys=[]
for data_key in code_list[0].keys():
    data_keys.append(data_key)
print(data_keys)   
#===================================================================================================
def write_to_xlsx():
    wb=Workbook()
    sheet=wb.active  #获取当前活跃的sheet,默认是第一个sheet
#     sheet=wb.create_sheet(title="sheet1",index=0)
#     sheet.cell(1,1).value = 0   #索引从1开始  
#     sheet.append([11,87])  #可以使用append插入一行数据  
    lg=len(code_list)
    
    sheet.append(data_keys)  #写入标题
    for i in range(lg):
        values=[]
        for ks,vs in code_list[i].items():
            if ks=='KlTm':
                values.append(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(vs)))
            else:
                values.append(vs)  #将一行cell拼成list
        sheet.append(values)  #按行写入
        
    wb.save(filename) 
#===================================================================================================
def read_from_xlsx():
    wb=load_workbook(filename)
    sheet=wb.active    #获取当前活跃的sheet,默认是第一个sheet
#     sheets=wb.sheetnames  #获取工作簿所有sheet名（list）
#     sheet=wb['sheetname'] # 指定sheet名
#     sheet1=wb[sheets[0]]   #获取第一个sheet

    rows=sheet.rows
    cols=sheet.columns
    for row in rows:
        line=[col.value for col in row]
        print(line)
        
    #通过坐标读取值  
#     cell_11 = booksheet.cell('A1').value  
#     cell_11 = booksheet.cell(row=1, column=1).value 
# print(sheet.max_column)  #获取最大列数
# print(sheet.max_row)  #获取最大行数
#===================================================================================================
# write_to_xlsx()
read_from_xlsx()
print(os.path.abspath('.'))
# os.system(filename)
